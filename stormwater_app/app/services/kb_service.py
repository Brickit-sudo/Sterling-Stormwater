"""
app/services/kb_service.py
Sterling Stormwater Knowledge Base — Excel reader service.

Reads assets/knowledge_base.xlsx and provides lookup functions
for write-up templates, photo captions, summary templates, and quick notes.

The workbook is cached in st.session_state["_kb_cache"] and reloaded
only when the file's mtime changes — no restart required after editing.
"""

from __future__ import annotations

import os
from pathlib import Path
from typing import Optional

import streamlit as st

KB_PATH = Path("assets/knowledge_base.xlsx")

_CACHE_KEY    = "_kb_cache"
_CACHE_MTIME  = "_kb_mtime"


# ─────────────────────────────────────────────────────────────────────────────
# Internal: load + cache
# ─────────────────────────────────────────────────────────────────────────────

def _load() -> dict:
    """Load and parse the knowledge base Excel file. Returns a dict of lists."""
    try:
        import openpyxl
    except ImportError:
        return {}

    if not KB_PATH.exists():
        return {}

    wb = openpyxl.load_workbook(KB_PATH, read_only=True, data_only=True)

    def _rows(sheet_name: str) -> list[list]:
        if sheet_name not in wb.sheetnames:
            return []
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        return [list(r) for r in rows if any(c is not None for c in r)]

    data = {
        "writeups":  _rows("WriteUps"),
        "captions":  _rows("PhotoCaptions"),
        "summaries": _rows("SummaryTemplates"),
        "notes":     _rows("QuickNotes"),
        "sites":     _rows("SiteProfiles"),
    }
    wb.close()
    return data


def _kb() -> dict:
    """Return cached KB data, reloading if the file has changed."""
    if not KB_PATH.exists():
        return {}

    mtime = KB_PATH.stat().st_mtime
    if (
        st.session_state.get(_CACHE_MTIME) != mtime
        or _CACHE_KEY not in st.session_state
    ):
        st.session_state[_CACHE_KEY]   = _load()
        st.session_state[_CACHE_MTIME] = mtime

    return st.session_state.get(_CACHE_KEY, {})


def kb_available() -> bool:
    """Return True if the knowledge base file exists."""
    return KB_PATH.exists()


# ─────────────────────────────────────────────────────────────────────────────
# Write-up lookup
# ─────────────────────────────────────────────────────────────────────────────

def get_writeup_options(
    system_type: str,
    field: str,
    condition: str = "ALL",
) -> list[dict]:
    """
    Return matching write-up templates for a given system type + field.

    Matches rows where:
      - system_type == system_type  OR  system_type == "ALL"
      - condition   == condition    OR  condition   == "ALL"
      - field       == field

    Returns list of {"label": str, "text": str} sorted: specific first, then ALL.
    """
    rows = _kb().get("writeups", [])
    specific, generic = [], []

    for row in rows:
        if len(row) < 5:
            continue
        st_val, cond_val, field_val, label, text = (row + [None]*5)[:5]
        if str(field_val or "").strip().lower() != field.lower():
            continue
        if str(cond_val or "").strip() not in (condition, "ALL"):
            continue

        entry = {"label": str(label or "").strip(), "text": str(text or "").strip()}
        if str(st_val or "").strip().upper() == "ALL":
            generic.append(entry)
        elif str(st_val or "").strip().lower() == system_type.lower():
            specific.append(entry)

    return specific + generic


def get_writeup_text(system_type: str, field: str, label: str) -> Optional[str]:
    """Return the text for a specific label."""
    for row in _kb().get("writeups", []):
        if len(row) < 5:
            continue
        st_val, _, field_val, lbl, text = (row + [None]*5)[:5]
        if (
            str(field_val or "").strip().lower() == field.lower()
            and str(lbl or "").strip() == label
            and str(st_val or "").strip().lower() in (system_type.lower(), "all")
        ):
            return str(text or "").strip()
    return None


# ─────────────────────────────────────────────────────────────────────────────
# Photo caption lookup
# ─────────────────────────────────────────────────────────────────────────────

def get_caption_options(system_type: str) -> list[dict]:
    """
    Return caption templates for a given system type (plus ALL generics).
    Returns list of {"label": str, "view": str, "caption": str}.
    """
    rows = _kb().get("captions", [])
    specific, generic = [], []

    for row in rows:
        if len(row) < 4:
            continue
        st_val, view, label, caption = (row + [None]*4)[:4]
        entry = {
            "label":   str(label   or "").strip(),
            "view":    str(view    or "").strip(),
            "caption": str(caption or "").strip(),
        }
        if str(st_val or "").strip().upper() == "ALL":
            generic.append(entry)
        elif str(st_val or "").strip().lower() == system_type.lower():
            specific.append(entry)

    return specific + generic


# ─────────────────────────────────────────────────────────────────────────────
# Summary template lookup
# ─────────────────────────────────────────────────────────────────────────────

def get_summary_options(report_type: str = "ALL") -> list[dict]:
    """
    Return summary templates for a given report type (plus ALL generics).
    Returns list of {"label": str, "text": str}.
    """
    rows = _kb().get("summaries", [])
    specific, generic = [], []

    for row in rows:
        if len(row) < 3:
            continue
        rt_val, label, text = (row + [None]*3)[:3]
        entry = {"label": str(label or "").strip(), "text": str(text or "").strip()}
        if str(rt_val or "").strip().upper() == "ALL":
            generic.append(entry)
        elif str(rt_val or "").strip().lower() == report_type.lower():
            specific.append(entry)

    return specific + generic


# ─────────────────────────────────────────────────────────────────────────────
# Quick notes lookup
# ─────────────────────────────────────────────────────────────────────────────

def get_quick_notes(category: str = "ALL") -> list[str]:
    """
    Return quick-note phrases for a given category (or all if category=="ALL").
    """
    rows = _kb().get("notes", [])
    results = []
    for row in rows:
        if len(row) < 2:
            continue
        cat, note = (row + [None]*2)[:2]
        if category.upper() == "ALL" or str(cat or "").strip().lower() == category.lower():
            results.append(str(note or "").strip())
    return [n for n in results if n]


def get_note_categories() -> list[str]:
    """Return all unique note categories."""
    rows = _kb().get("notes", [])
    seen, cats = set(), []
    for row in rows:
        if len(row) < 1:
            continue
        cat = str(row[0] or "").strip()
        if cat and cat not in seen:
            seen.add(cat)
            cats.append(cat)
    return cats


# ─────────────────────────────────────────────────────────────────────────────
# Site profiles lookup
# ─────────────────────────────────────────────────────────────────────────────

def get_site_profiles() -> list[dict]:
    """Return all site profiles."""
    rows = _kb().get("sites", [])
    results = []
    for row in rows:
        if len(row) < 1:
            continue
        site, client, addr, systems, notes = (list(row) + [None]*5)[:5]
        results.append({
            "site_name":   str(site    or "").strip(),
            "client_name": str(client  or "").strip(),
            "address":     str(addr    or "").strip(),
            "systems":     [s.strip() for s in str(systems or "").split(",") if s.strip()],
            "notes":       str(notes   or "").strip(),
        })
    return [r for r in results if r["site_name"]]


def get_site_profile(site_name: str) -> Optional[dict]:
    """Return a specific site profile by name."""
    for p in get_site_profiles():
        if p["site_name"].lower() == site_name.lower():
            return p
    return None
