"""
create_knowledge_base.py
Generate the Sterling Stormwater knowledge base Excel file.

Run once:  python create_knowledge_base.py
Output:    assets/knowledge_base.xlsx

Sheets:
  1. WriteUps          — findings / recommendations / maintenance / post-service templates
  2. PhotoCaptions      — caption templates by system type + view
  3. SummaryTemplates   — overall summary paragraphs
  4. QuickNotes         — field note quick-insert phrases
  5. SiteProfiles       — reusable site configurations
"""

from pathlib import Path
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

OUT_PATH = Path("assets/knowledge_base.xlsx")

# ── Brand colors ──────────────────────────────────────────────────────────────
NAVY   = "0B2A3C"
GREEN  = "1AB738"
DARK   = "06141C"
LIGHT  = "F1F5F9"
MID    = "103447"
MUTED  = "6B7A8A"

def _hdr(ws, row: int, cols: list[str]):
    """Write header row with navy fill + bold white text."""
    for col, label in enumerate(cols, 1):
        cell = ws.cell(row=row, column=col, value=label)
        cell.font      = Font(bold=True, color=LIGHT, size=11, name="Calibri")
        cell.fill      = PatternFill("solid", fgColor=NAVY)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = Border(
            bottom=Side(style="thin", color=GREEN),
            right=Side(style="hair", color=MID),
        )

def _row(ws, row: int, values: list, wrap=True):
    """Write a data row."""
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.alignment = Alignment(vertical="top", wrap_text=wrap)
        cell.border    = Border(
            bottom=Side(style="hair", color="C9D3DD"),
            right=Side(style="hair",  color="E3E8EE"),
        )
        # Zebra stripe
        if row % 2 == 0:
            cell.fill = PatternFill("solid", fgColor="F0F4F8")

def _col_widths(ws, widths: list[int]):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def _freeze(ws, cell="A2"):
    ws.freeze_panes = cell

def _tab_color(ws, hex_color: str):
    ws.sheet_properties.tabColor = hex_color

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# SHEET 1 — WriteUps
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
WRITEUP_DATA = [
    # (system_type, condition, field, label, text)

    # ── Bioretention Cell ──────────────────────────────────────────────────────
    ("Bioretention Cell", "Good", "findings", "Standard - Good",
     "The Bioretention Cell was inspected and found to be in good condition at the time of inspection. "
     "The inlet curb cut and/or inlet pipe was clear of obstructions with no evidence of erosion or clogging. "
     "The surface media appeared in good condition with appropriate vegetation coverage and no significant sediment accumulation. "
     "The overflow structure was intact and unobstructed. The underdrain cleanout cap was in place. "
     "No structural deficiencies were observed."),

    ("Bioretention Cell", "Fair", "findings", "Standard - Fair",
     "The Bioretention Cell was inspected and found to be in fair condition at the time of inspection. "
     "Minor sediment accumulation was observed at the inlet area. Vegetation coverage was adequate, though some bare areas were noted. "
     "The overflow structure was functional. The underdrain cleanout cap was present. "
     "Routine maintenance is recommended to maintain optimal performance."),

    ("Bioretention Cell", "Poor", "findings", "Standard - Poor",
     "The Bioretention Cell was inspected and found to be in poor condition at the time of inspection. "
     "Significant sediment accumulation was observed at the inlet and within the cell, reducing infiltration capacity. "
     "Vegetation was sparse or absent in areas, and evidence of erosion was noted on the side slopes. "
     "The overflow structure showed signs of obstruction. Immediate corrective maintenance is recommended."),

    ("Bioretention Cell", "Good", "recommendations", "No Action Required",
     "No corrective maintenance is recommended at this time. The Bioretention Cell is operating as designed. "
     "Routine maintenance should continue per the approved Operation and Maintenance Plan."),

    ("Bioretention Cell", "Fair", "recommendations", "Routine Maintenance",
     "The following maintenance is recommended: (1) Remove accumulated sediment at the inlet area. "
     "(2) Replant bare vegetation areas with approved plant species. "
     "(3) Inspect and clear the underdrain cleanout. "
     "(4) Continue monitoring per the approved O&M Plan."),

    ("Bioretention Cell", "Poor", "recommendations", "Corrective Maintenance",
     "Corrective maintenance is recommended as follows: (1) Remove all accumulated sediment from the inlet and cell surface. "
     "(2) Scarify and aerate compacted surface media to restore infiltration. "
     "(3) Replant disturbed vegetation areas. "
     "(4) Clear overflow structure of all obstructions. "
     "(5) Evaluate need for media replacement if infiltration does not recover after maintenance."),

    ("Bioretention Cell", "ALL", "maintenance_performed", "Standard Maintenance",
     "The following maintenance was performed at the Bioretention Cell: "
     "Inlet area was cleared of accumulated sediment and debris. "
     "Surface of the bioretention media was raked and aerated where compaction was observed. "
     "Overflow structure was inspected and cleared. "
     "Underdrain cleanout cap was inspected and secured. "
     "Debris and litter were removed from the cell and surrounding area."),

    ("Bioretention Cell", "ALL", "post_service_condition", "Post-Service Good",
     "Following maintenance, the Bioretention Cell was found to be in good condition. "
     "Inlet flow path was clear and unobstructed. Surface media was in acceptable condition. "
     "No outstanding follow-up items were identified at this time."),

    # ── Catch Basin / Inlet ────────────────────────────────────────────────────
    ("Catch Basin / Inlet", "Good", "findings", "Standard - Good",
     "The Catch Basin was inspected and found to be in good condition at the time of inspection. "
     "The grate and frame were intact and free of debris. The sump was clear with minimal sediment accumulation. "
     "The outlet pipe was unobstructed and showed no signs of structural deficiency. No odors were detected."),

    ("Catch Basin / Inlet", "Fair", "findings", "Standard - Fair",
     "The Catch Basin was inspected and found to be in fair condition at the time of inspection. "
     "Moderate sediment accumulation was observed in the sump. The grate was partially obstructed by debris. "
     "The outlet pipe appeared functional. Cleaning is recommended at the next scheduled maintenance visit."),

    ("Catch Basin / Inlet", "Poor", "findings", "Standard - Poor",
     "The Catch Basin was inspected and found to be in poor condition at the time of inspection. "
     "Heavy sediment accumulation was observed in the sump, with sediment depth exceeding 4 inches. "
     "The grate was significantly obstructed. The outlet pipe showed evidence of restriction. Immediate cleaning is recommended."),

    ("Catch Basin / Inlet", "ALL", "maintenance_performed", "Standard Cleaning",
     "The Catch Basin was cleaned using a vacuum truck. Accumulated sediment and debris were removed from the sump. "
     "The grate and frame were cleared of all debris. The outlet pipe was inspected and confirmed unobstructed. "
     "The catch basin was left in clean, functional condition."),

    ("Catch Basin / Inlet", "Good", "recommendations", "No Action Required",
     "No corrective maintenance is recommended at this time. Continue routine inspection and cleaning per the approved O&M Plan."),

    # ── Underdrain Soil Filter ─────────────────────────────────────────────────
    ("Underdrain Soil Filter", "Good", "findings", "Standard - Good",
     "The Underdrained Soil Filter (USF) was inspected and found to be in good condition at the time of inspection. "
     "The inlet structure was clear and functioning properly. The filter media surface appeared in good condition "
     "with appropriate vegetation and minimal sediment accumulation. The underdrain outlet was unobstructed and "
     "showed no signs of structural damage. No bypass or short-circuiting was observed."),

    ("Underdrain Soil Filter", "Fair", "findings", "Standard - Fair",
     "The Underdrained Soil Filter (USF) was inspected and found to be in fair condition. "
     "Some sediment accumulation was observed at the inlet. The filter media appeared partially compacted in areas. "
     "The underdrain outlet was functional. Routine maintenance is recommended."),

    ("Underdrain Soil Filter", "Poor", "findings", "Standard - Poor",
     "The Underdrained Soil Filter (USF) was inspected and found to be in poor condition. "
     "Heavy sediment loading was observed on the filter surface, significantly reducing infiltration capacity. "
     "Evidence of ponding and short-circuiting was noted. Corrective maintenance including media cleaning or replacement is recommended."),

    ("Underdrain Soil Filter", "ALL", "maintenance_performed", "Standard Maintenance",
     "The following maintenance was performed at the Underdrained Soil Filter: "
     "Inlet structure was cleared of all accumulated sediment and debris. "
     "Filter media surface was raked and aerated. "
     "Underdrain outlet pipe was inspected and confirmed unobstructed. "
     "All debris and litter were removed from the facility and surrounding area."),

    # ── Wet Pond ───────────────────────────────────────────────────────────────
    ("Wet Pond", "Good", "findings", "Standard - Good",
     "The Wet Pond was inspected and found to be in good condition at the time of inspection. "
     "The primary outlet riser and barrel were intact and unobstructed. The emergency spillway was clear and functional. "
     "The forebay contained minimal sediment accumulation. Embankment slopes were stable with adequate vegetative cover. "
     "Water quality appeared acceptable with no evidence of excessive algae or floating debris."),

    ("Wet Pond", "Fair", "findings", "Standard - Fair",
     "The Wet Pond was inspected and found to be in fair condition. "
     "Moderate sediment accumulation was observed in the forebay. Some bare areas were noted on the embankment slopes. "
     "The primary outlet was functional. Routine maintenance including forebay cleanout is recommended."),

    ("Wet Pond", "Poor", "findings", "Standard - Poor",
     "The Wet Pond was inspected and found to be in poor condition. "
     "Significant sediment accumulation has reduced the forebay storage capacity. Embankment erosion was observed. "
     "Evidence of short-circuiting between inlet and outlet was noted. "
     "The outlet riser showed signs of obstruction. Corrective dredging and embankment repair are recommended."),

    # ── Generic / ALL ──────────────────────────────────────────────────────────
    ("ALL", "Good", "findings", "Generic - Good Condition",
     "The stormwater management system was inspected and found to be in good condition at the time of inspection. "
     "All components were functioning as designed with no significant deficiencies observed."),

    ("ALL", "Fair", "findings", "Generic - Fair Condition",
     "The stormwater management system was inspected and found to be in fair condition at the time of inspection. "
     "Minor deficiencies were observed. Routine maintenance is recommended to prevent further deterioration."),

    ("ALL", "Poor", "findings", "Generic - Poor Condition",
     "The stormwater management system was inspected and found to be in poor condition at the time of inspection. "
     "Significant deficiencies were observed that require corrective maintenance to restore proper function."),

    ("ALL", "Good", "recommendations", "No Action Required",
     "No corrective maintenance is recommended at this time. The system is operating as designed. "
     "Continue routine maintenance per the approved O&M Plan."),

    ("ALL", "Fair", "recommendations", "Routine Maintenance Recommended",
     "Routine maintenance is recommended at the next scheduled service visit. "
     "Continue monitoring per the approved Operation and Maintenance Plan."),

    ("ALL", "Poor", "recommendations", "Corrective Action Required",
     "Corrective maintenance is recommended as soon as practicable. "
     "Failure to address observed deficiencies may result in reduced system performance and potential regulatory non-compliance."),
]

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# SHEET 2 — PhotoCaptions
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
CAPTION_DATA = [
    # (system_type, view, label, caption_template)
    # Use {system_id} and {system_name} as placeholders
    ("Bioretention Cell",    "Overall View",       "Standard Overall",       "{system_name} – Overall View"),
    ("Bioretention Cell",    "Inlet / Curb Cut",   "Inlet Condition",        "{system_name} – View of Inlet Area"),
    ("Bioretention Cell",    "Overflow Structure", "Overflow",               "{system_name} – View of Overflow Structure"),
    ("Bioretention Cell",    "Surface Media",      "Media Surface",          "{system_name} – View of Surface Media"),
    ("Bioretention Cell",    "Vegetation",         "Vegetation",             "{system_name} – View of Vegetation"),
    ("Bioretention Cell",    "Outlet",             "Underdrain Outlet",      "{system_name} – View of Underdrain Outlet Pipe"),
    ("Catch Basin / Inlet",  "Overall View",       "Standard Overall",       "{system_name} – Overall View"),
    ("Catch Basin / Inlet",  "Grate / Frame",      "Grate Condition",        "{system_name} – View of Grate/Frame"),
    ("Catch Basin / Inlet",  "Sump",               "Sump Condition",         "{system_name} – View Inside Sump"),
    ("Catch Basin / Inlet",  "Sediment Accumulation", "Sediment",            "{system_name} – Sediment Accumulation"),
    ("Underdrain Soil Filter", "Overall View",     "Standard Overall",       "{system_name} – Overall View"),
    ("Underdrain Soil Filter", "Inlet Structure",  "Inlet",                  "{system_name} – View of Inlet Structure"),
    ("Underdrain Soil Filter", "Surface Media",    "Filter Surface",         "{system_name} – View of Filter Media Surface"),
    ("Underdrain Soil Filter", "Outlet Structure", "Outlet/Underdrain",      "{system_name} – View of Underdrain Outlet"),
    ("Wet Pond",             "Overall View",       "Standard Overall",       "{system_name} – Overall View"),
    ("Wet Pond",             "Primary Outlet / Riser", "Outlet Riser",       "{system_name} – View of Primary Outlet Riser"),
    ("Wet Pond",             "Forebay",            "Forebay",                "{system_name} – View of Forebay"),
    ("Wet Pond",             "Embankment",         "Embankment",             "{system_name} – View of Embankment"),
    ("Wet Pond",             "Emergency Spillway", "Spillway",               "{system_name} – View of Emergency Spillway"),
    ("ALL",                  "Overall View",       "Generic Overall",        "{system_name} – Overall View"),
    ("ALL",                  "Inlet",              "Generic Inlet",          "{system_name} – View of Inlet"),
    ("ALL",                  "Outlet",             "Generic Outlet",         "{system_name} – View of Outlet"),
    ("ALL",                  "After Maintenance",  "Post-Maintenance",       "{system_name} – After Maintenance"),
]

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# SHEET 3 — SummaryTemplates
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
SUMMARY_DATA = [
    # (report_type, label, text)
    ("Inspection", "Standard Inspection Opening",
     "An inspection of the above referenced stormwater management systems was performed on [inspection date]. "
     "The results of the inspection revealed the following conditions for each facility:"),

    ("Inspection", "All Good Condition",
     "An inspection of the above referenced stormwater management systems was performed on [inspection date]. "
     "All systems were found to be in good condition and operating as designed. "
     "No corrective action is recommended at this time. Routine maintenance should continue per the approved O&M Plan."),

    ("Maintenance", "Standard Maintenance Opening",
     "Maintenance services were performed on the above referenced stormwater management systems on [service date]. "
     "The following is a summary of conditions observed and work performed at each facility:"),

    ("Maintenance", "All Systems Serviced",
     "Maintenance services were performed on the above referenced stormwater management systems on [service date]. "
     "All systems were cleaned and serviced per the approved Operation and Maintenance Plan. "
     "Following maintenance, all systems were found to be in good condition."),

    ("Inspection and Maintenance", "Combined Opening",
     "An inspection and maintenance visit was performed for the above referenced stormwater management systems on [service date]. "
     "The following is a summary of conditions observed prior to service and maintenance activities performed at each facility:"),

    ("ALL", "Regulatory Compliance Note",
     "This inspection was performed in accordance with the requirements of the approved Operation and Maintenance Agreement and "
     "applicable local stormwater regulations. All findings and recommendations are summarized herein."),
]

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# SHEET 4 — QuickNotes
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
QUICKNOTE_DATA = [
    # (category, note)
    ("Sediment",   ">4\" Sediment"),
    ("Sediment",   "Heavy Sediment"),
    ("Sediment",   "Moderate Sediment"),
    ("Sediment",   "Light Sediment"),
    ("Obstruction","Fully Obstructed"),
    ("Obstruction","Partially Obstructed"),
    ("Obstruction","Debris Present"),
    ("Water",      "Standing Water"),
    ("Water",      "Sheen on Water"),
    ("Water",      "Discoloration Observed"),
    ("Vegetation", "Excess Vegetation"),
    ("Vegetation", "Sparse Vegetation"),
    ("Vegetation", "Invasive Species Present"),
    ("Vegetation", "Mowing Required"),
    ("Structural", "Structural Damage"),
    ("Structural", "Crack Observed"),
    ("Structural", "Detached Hood"),
    ("Structural", "Missing Hardware"),
    ("Structural", "Missing Cap"),
    ("Erosion",    "Erosion Present"),
    ("Erosion",    "Minor Erosion"),
    ("Erosion",    "Significant Erosion"),
    ("Erosion",    "Animal Burrow"),
    ("Severity",   "Minimal"),
    ("Severity",   "Light"),
    ("Severity",   "Moderate"),
    ("Severity",   "Heavy"),
    ("Severity",   "Significant"),
    ("Severity",   "Excessive"),
    ("General",    "No Issues Observed"),
    ("General",    "Operating As Designed"),
    ("General",    "Requires Follow-Up"),
    ("General",    "See Photos"),
]

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# SHEET 5 — SiteProfiles
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
SITE_DATA = [
    # (site_name, client_name, address, systems_csv, notes)
    ("Example Site A", "ABC Property Management", "123 Main St, Anytown MD 20001",
     "USF-1, USF-2, CB-1, CB-2, CB-3", "Annual inspection required by permit. Contact: John Smith 301-555-0100"),
    ("Example Site B", "XYZ Corporation", "456 Commerce Dr, Rockville MD 20850",
     "BR-1, BR-2, WP-1", "Quarterly maintenance. HOA managed property."),
]


def build():
    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    # ── Sheet 1: WriteUps ─────────────────────────────────────────────────────
    ws1 = wb.create_sheet("WriteUps")
    _tab_color(ws1, GREEN)
    _hdr(ws1, 1, ["System Type", "Condition", "Field", "Label", "Text"])
    for i, row in enumerate(WRITEUP_DATA, 2):
        _row(ws1, i, list(row))
    ws1.row_dimensions[1].height = 22
    for i in range(2, len(WRITEUP_DATA) + 2):
        ws1.row_dimensions[i].height = 60
    _col_widths(ws1, [28, 12, 24, 28, 80])
    _freeze(ws1, "A2")
    ws1.auto_filter.ref = f"A1:E{len(WRITEUP_DATA)+1}"

    # ── Sheet 2: PhotoCaptions ────────────────────────────────────────────────
    ws2 = wb.create_sheet("PhotoCaptions")
    _tab_color(ws2, "38BDF8")
    _hdr(ws2, 1, ["System Type", "View / Component", "Label", "Caption Template"])
    for i, row in enumerate(CAPTION_DATA, 2):
        _row(ws2, i, list(row))
    ws2.row_dimensions[1].height = 22
    for i in range(2, len(CAPTION_DATA) + 2):
        ws2.row_dimensions[i].height = 30
    _col_widths(ws2, [28, 24, 26, 55])
    _freeze(ws2, "A2")
    ws2.auto_filter.ref = f"A1:D{len(CAPTION_DATA)+1}"

    # ── Sheet 3: SummaryTemplates ─────────────────────────────────────────────
    ws3 = wb.create_sheet("SummaryTemplates")
    _tab_color(ws3, "F59E0B")
    _hdr(ws3, 1, ["Report Type", "Label", "Text"])
    for i, row in enumerate(SUMMARY_DATA, 2):
        _row(ws3, i, list(row))
    ws3.row_dimensions[1].height = 22
    for i in range(2, len(SUMMARY_DATA) + 2):
        ws3.row_dimensions[i].height = 50
    _col_widths(ws3, [28, 30, 90])
    _freeze(ws3, "A2")
    ws3.auto_filter.ref = f"A1:C{len(SUMMARY_DATA)+1}"

    # ── Sheet 4: QuickNotes ───────────────────────────────────────────────────
    ws4 = wb.create_sheet("QuickNotes")
    _tab_color(ws4, MUTED)
    _hdr(ws4, 1, ["Category", "Note / Phrase"])
    for i, row in enumerate(QUICKNOTE_DATA, 2):
        _row(ws4, i, list(row))
    ws4.row_dimensions[1].height = 22
    for i in range(2, len(QUICKNOTE_DATA) + 2):
        ws4.row_dimensions[i].height = 18
    _col_widths(ws4, [18, 40])
    _freeze(ws4, "A2")
    ws4.auto_filter.ref = f"A1:B{len(QUICKNOTE_DATA)+1}"

    # ── Sheet 5: SiteProfiles ─────────────────────────────────────────────────
    ws5 = wb.create_sheet("SiteProfiles")
    _tab_color(ws5, NAVY)
    _hdr(ws5, 1, ["Site Name", "Client Name", "Address", "Systems (comma-separated)", "Notes"])
    for i, row in enumerate(SITE_DATA, 2):
        _row(ws5, i, list(row))
    ws5.row_dimensions[1].height = 22
    for i in range(2, len(SITE_DATA) + 2):
        ws5.row_dimensions[i].height = 30
    _col_widths(ws5, [28, 28, 35, 35, 55])
    _freeze(ws5, "A2")

    wb.save(OUT_PATH)
    print(f"Knowledge base saved to: {OUT_PATH.resolve()}")
    print(f"    WriteUps:         {len(WRITEUP_DATA)} templates")
    print(f"    PhotoCaptions:    {len(CAPTION_DATA)} templates")
    print(f"    SummaryTemplates: {len(SUMMARY_DATA)} templates")
    print(f"    QuickNotes:       {len(QUICKNOTE_DATA)} phrases")
    print(f"    SiteProfiles:     {len(SITE_DATA)} example sites")
    print()
    print("Open assets/knowledge_base.xlsx to add your own templates.")
    print("The app reads it live — no restart needed after editing.")


if __name__ == "__main__":
    build()
